from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font, Alignment


#SOME USEFUL FUNCTIONS IN BEAUTIFULSOUP
#-----------------------------------------------#
# find(tag, attributes, recursive, text, keywords)
# findAll(tag, attributes, recursive, text, limit, keywords)

#Tags: find("h1","h2","h3", etc.)
#Attributes: find("span", {"class":{"green","red"}})
#Text: nameList = Objfind(text="the prince")
#Limit = find with limit of 1
#keyword: allText = Obj.find(id="title",class="text")


url = 'https://www.coingecko.com/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

# print(soup.title)

# --------------------------------------------------

# setup workbook
wb = xl.Workbook()
ws = wb.active
ws.title = 'Crypto Report'

# assign title & headers to new file
ws['A1'] = 'Cryptocurrency Prices by Market Cap'
ws['A2'] = 'based on ' + url
ws['A4'] = 'No.'
ws['B4'] = 'Name'
ws['C4'] = 'Code'
ws['D4'] = 'Current Price'
ws['E4'] = '% Change in the Last 24 Hours'
ws['F4'] = 'Price Based on % Change'

ws.merge_cells('A1:F1')
ws.merge_cells('A2:F2')

crypto_rows = soup.findAll("tr")

for x in range(1,6):
    # webscrapping
    td = crypto_rows[x].findAll('td')
    span = crypto_rows[x].findAll('span')

    no = td[1].text.strip('\n')
    name = span[0].text.strip('\n')
    code = span[1].text.strip('\n')
    current_price = float(td[3].text.replace(",","").replace("$","").strip('\n'))
    percent_change = float(td[5].text.replace("%","").strip('\n'))

    percent_change_in_hundreds = percent_change / 100
    price_change = current_price * percent_change_in_hundreds
    
    # write sheet
    ws['A' + str(x+4)] = no
    ws['B' + str(x+4)] = name
    ws['C' + str(x+4)] = code
    ws['D' + str(x+4)] = current_price
    # ws['D' + str(x+3)] = '$' + str(format(current_price, '.2f'))
    ws['E' + str(x+4)] = str(percent_change) + '%'
    ws['F' + str(x+4)] = price_change
    # ws['F' + str(x+3)] = '$'+ str(format(price_change, '.2f'))

    # --------------------------------------------------

    # twilio text alert
    import keys
    from twilio.rest import Client

    client = Client(keys.accountSID, keys.authToken)

    # enter your own personal numbers
    TwilioNumber = '+12544525613'
    mycellphone = '+15126941668'

    if name == 'Bitcoin':
        if price_change >= 5:
            crypto_alert = f"The cryptocurrency {name} has increased or decreased within $5 of its current value with a percent change of {str(percent_change) + '%'} and a corresponding price of {'$'+ str(format(price_change, '.2f'))}"
            # uncomment to send text message alerts
            # textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber, body=crypto_alert)
            print(crypto_alert)

    if name == 'Ethereum':
        if price_change >= 5:
            crypto_alert = f"The cryptocurrency {name} has increased or decreased within $5 of its current value with a percent change of {str(percent_change) + '%'} and a corresponding price of {'$'+ str(format(price_change, '.2f'))}"
            # uncomment to send text message alerts
            # textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber, body=crypto_alert)
            print(crypto_alert)

# --------------------------------------------------

# formatting
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 31
ws.column_dimensions['F'].width = 26

header_font = Font(size=24, bold=True, color='ffffcc00')
category_font = Font(size=12, bold=True, italic=True)

for cell in ws[1:1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for cell in ws[2:2]:
    cell.alignment = Alignment(horizontal='center')

for cell in ws[4:4]:
    cell.font = category_font
    cell.alignment = Alignment(horizontal='center')

for cell in ws['D:D']:
    cell.number_format = u'"$ "#,##0.00'
    cell.alignment = Alignment(horizontal='center')

for cell in ws['E:E']:
    cell.alignment = Alignment(horizontal='center')

for cell in ws['F:F']:
    cell.number_format = u'"$ "#,##0.00'
    cell.alignment = Alignment(horizontal='center')

# # save workbook
wb.save("CryptoReport.xlsx")